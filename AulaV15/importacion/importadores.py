import csv
from pathlib import Path
from .interfaces import IImportador


class CSVImportador(IImportador):
    """Lee archivos .csv usando encabezados en la primera fila."""

    def leer(self, ruta_archivo: str) -> list[dict]:
        with open(ruta_archivo, "r", encoding="utf-8-sig", newline="") as archivo:
            return [dict(fila) for fila in csv.DictReader(archivo)]


class ExcelImportador(IImportador):
    """Lee archivos .xlsx usando openpyxl sin acoplar la lógica principal a Excel."""

    def leer(self, ruta_archivo: str) -> list[dict]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("Para importar Excel instale openpyxl: pip install openpyxl") from exc

        wb = load_workbook(ruta_archivo, read_only=True, data_only=True)
        hoja = wb.active
        filas = list(hoja.iter_rows(values_only=True))
        if not filas:
            return []
        encabezados = [str(c).strip() if c is not None else "" for c in filas[0]]
        datos: list[dict] = []
        for fila in filas[1:]:
            datos.append({encabezados[i]: fila[i] if i < len(fila) else "" for i in range(len(encabezados))})
        return datos
